# -*- coding: utf-8 -*-
# © 2015-2017 Andrei Levin - Didotech srl (www.didotech.com)

from openerp import _
from openerp import exceptions


def import_sheet(filename, content):
    name, file_type = filename.rsplit('.', 1)

    if file_type in ('xls', 'xlsb'):
        # "Excel"
        import xlrd

        for encoding in ('utf-8', 'latin-1', 'cp1252'):
            try:
                book = xlrd.open_workbook(file_contents=content, encoding_override=encoding)
                break
            except UnicodeDecodeError:
                pass
        else:
            raise exceptions.Warning(_('Error: Unknown encoding'))

        table = []
        sh = book.sheet_by_index(0)

        for rx in range(sh.nrows):
            row = []
            for cx in range(sh.ncols):
                row.append(sh.cell(rowx=rx, colx=cx).value)
            table.append(row)
        number_of_lines = sh.nrows
    elif file_type in ('xlsx', 'xlsm', 'xltx', 'xltm'):
        import openpyxl
        import StringIO

        ## Create virtual File:
        virtual_file = StringIO.StringIO(content)
        book = openpyxl.load_workbook(virtual_file, read_only=True)

        table = []
        sh = book.worksheets[0]
        max_column = sh.max_column

        for rx in range(1, sh.max_row + 1):
            row = []
            for cx in range(1, max_column + 1):
                if rx == 1 and not sh.cell(row=rx, column=cx).value:
                    max_column = cx - 1
                    break

                row.append(sh.cell(row=rx, column=cx).value)
            table.append(row)
        number_of_lines = sh.max_row
    elif file_type in ('ods', ):
        # "OpenOffice"
        from openerp.addons.core_extended.odf_to_array import ODSReader
        import StringIO

        ## Create virtual File:
        virtual_file = StringIO.StringIO(content)

        book = ODSReader(virtual_file)
        table = book.sheet_by_index(0)
        number_of_lines = len(table)
    elif file_type == 'csv':
        # "CSV"
        import csv
        import StringIO

        def unicode_csv_reader(unicode_csv_data, dialect=csv.excel, **kwargs):
            # csv.py doesn't do Unicode; encode temporarily as UTF-8:
            csv_reader = csv.reader(utf_8_encoder(unicode_csv_data),
                                    dialect=dialect, **kwargs)
            for row in csv_reader:
                # decode UTF-8 back to Unicode, cell by cell:
                yield [unicode(cell, 'utf-8') for cell in row]

        def utf_8_encoder(unicode_csv_data):
            for line in unicode_csv_data:
                yield line.encode('utf-8')

        ## Create virtual File:
        virtual_file = StringIO.StringIO(content)
        virtual_file_utf8 = StringIO.StringIO(content.decode('utf-8'))

        ## Process CSV file:
        sample = virtual_file.read(512)
        virtual_file.seek(0)
        dialect = csv.Sniffer().sniff(sample)

        table = csv.reader(virtual_file, dialect)
        table_utf8 = unicode_csv_reader(virtual_file_utf8, dialect)

        # self.table is an object of type '_csv.reader' and has no len() method
        number_of_lines = sum(1 for row in table)
        virtual_file.seek(0)
    else:
        raise exceptions.Warning(_('Error: Unknown file extension'))

    return table_utf8, number_of_lines
